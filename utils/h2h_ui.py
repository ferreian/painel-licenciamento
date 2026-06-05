"""
utils/h2h_ui.py — Helpers visuais de H2H reaproveitados do painel JAUM DTC.
Classificação por % de vitórias, tabela AgGrid estilizada, export Excel
colorido e donut de classes. Mantém a mesma identidade visual do painel original.
"""

import io

import pandas as pd
import plotly.graph_objects as go
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, JsCode


# ── Comunicação de dados (estilo Schwabish) ──────────────────────────────────
def insight_card(pergunta: str, resposta_html: str):
    """Callout em pergunta-e-resposta destacando a mensagem principal do gráfico."""
    st.markdown(
        f'<div style="border-left:4px solid #27AE60;background:#E9F7EF;border-radius:8px;'
        f'padding:12px 16px;margin:4px 0 20px;">'
        f'<p style="font-size:13px;font-weight:700;color:#1E8449;text-transform:uppercase;'
        f'letter-spacing:0.04em;margin:0 0 5px;">{pergunta}</p>'
        f'<p style="font-size:17px;color:#1A1A1A;line-height:1.6;margin:0;">{resposta_html}</p>'
        f'</div>', unsafe_allow_html=True)


def como_ler(markdown_texto: str, label: str = "ℹ️ Como ler este gráfico"):
    """Popover padrão de 'como ler', ao lado do título do gráfico."""
    with st.popover(label, use_container_width=False):
        st.markdown(markdown_texto)


# Legenda das classes — em markdown (para modais) e em chips (para a seção Leitura)
LEGENDA_CLASSES_MD = (
    "**Faixas de % de vitórias (locais em que o material venceu):**\n\n"
    "- 🟩 **Alta Performance** — vence em **> 75%** dos locais\n"
    "- 🟦 **Superior** — **56–75%**\n"
    "- 🟨 **Competitivo** — **46–55%** (equilíbrio)\n"
    "- 🟥 **Restrito** — **≤ 45%** (atenção ao posicionamento)\n"
)


def legenda_classes_chips():
    """Renderiza a legenda das 4 classes como chips coloridos (para a seção Leitura)."""
    chips = [
        ("Alta Performance", "#90EE90", "> 75%"),
        ("Superior",         "#87CEFF", "56–75%"),
        ("Competitivo",      "#FFFF00", "46–55%"),
        ("Restrito",         "#FF0000", "≤ 45%"),
    ]
    html = '<div style="display:flex;flex-wrap:wrap;gap:8px;margin:2px 0 16px;">'
    for nome, cor, faixa in chips:
        html += (
            '<span style="display:inline-flex;align-items:center;gap:7px;background:#FFFFFF;'
            'border:1px solid #E5E7EB;border-radius:20px;padding:6px 13px;font-size:14px;">'
            f'<span style="width:13px;height:13px;border-radius:3px;background:{cor};'
            'display:inline-block;border:1px solid rgba(0,0,0,0.1);"></span>'
            f'<b style="color:#1A1A1A;">{nome}</b>'
            f'<span style="color:#6B7280;">{faixa}</span></span>')
    html += "</div>"
    st.markdown(html, unsafe_allow_html=True)


def chart_pergunta(texto: str):
    """Pergunta que o gráfico responde — exibida acima do gráfico, sem caixa."""
    st.markdown(
        f'<p style="font-size:20px;font-weight:700;color:#1A1A1A;margin:2px 0 6px;'
        f'line-height:1.3;">{texto}</p>', unsafe_allow_html=True)


def chart_resposta(html: str):
    """Resposta/insight — exibida logo após a pergunta, como texto solto (sem caixa)."""
    st.markdown(
        f'<p style="font-size:16px;color:#1A1A1A;line-height:1.6;margin:0 0 12px;">'
        f'{html}</p>', unsafe_allow_html=True)


# ── Régua de classificação (idêntica ao painel) ──────────────────────────────
def classificar_h2h(pct: float) -> tuple[str, str]:
    """Retorna (label, cor_fundo_hex) pelo % de vitórias (WR%)."""
    if pd.isna(pct):
        return "—", "#F3F4F6"
    if pct <= 45:
        return "Restrito", "#FF0000"
    elif pct <= 55:
        return "Competitivo", "#FFFF00"
    elif pct <= 75:
        return "Superior", "#87CEFF"
    else:
        return "Alta Performance", "#90EE90"


COR_FUNDO = {
    "Alta Performance": "#90EE90",
    "Superior":         "#87CEFF",
    "Competitivo":      "#FFFF00",
    "Restrito":         "#FF0000",
    "—":                "#F3F4F6",
}
COR_TEXTO = {
    "Alta Performance": "#1A1A1A",
    "Superior":         "#1A1A1A",
    "Competitivo":      "#1A1A1A",
    "Restrito":         "#FFFFFF",
    "—":                "#6B7280",
}


# ── Tabela AgGrid padrão ──────────────────────────────────────────────────────
def ag_table_h2h(df: pd.DataFrame, height: int = 480, key: str = "ag"):
    classe_style = JsCode("""
    function(params) {
        const v = params.value;
        if (v === 'Alta Performance') return {'backgroundColor':'#90EE90','color':'#1A1A1A','fontWeight':'700','textAlign':'center'};
        if (v === 'Superior')         return {'backgroundColor':'#87CEFF','color':'#1A1A1A','fontWeight':'700','textAlign':'center'};
        if (v === 'Competitivo')      return {'backgroundColor':'#FFFF00','color':'#1A1A1A','fontWeight':'700','textAlign':'center'};
        if (v === 'Restrito')         return {'backgroundColor':'#FF0000','color':'#FFFFFF','fontWeight':'700','textAlign':'center'};
        return {};
    }""")

    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(
        resizable=True, sortable=True, filter=True, suppressMenu=False,
        menuTabs=["generalMenuTab", "filterMenuTab", "columnsMenuTab"],
        cellStyle={"fontSize": "13px", "color": "#000000",
                   "fontFamily": "Helvetica Neue, sans-serif"},
    )
    if "Classe" in df.columns:
        gb.configure_column("Classe", cellStyle=classe_style, minWidth=140)
    # Headers coloridos para colunas cujo nome é uma classe (ex.: tabela da pág. 2)
    HDR_CLS = {"Alta Performance": "hdr-alta", "Superior": "hdr-sup",
               "Competitivo": "hdr-comp", "Restrito": "hdr-restr"}
    for _col, _cls in HDR_CLS.items():
        if _col in df.columns:
            gb.configure_column(_col, headerClass=_cls)
    gb.configure_grid_options(
        headerHeight=36, rowHeight=32, domLayout="normal",
        suppressMenuHide=True, suppressColumnVirtualisation=True,
        enableRangeSelection=True,
    )
    go_opts = gb.build()
    go_opts["onFirstDataRendered"] = JsCode(
        "function(params){ params.api.sizeColumnsToFit(); }"
    )
    AgGrid(
        df, gridOptions=go_opts, height=height,
        update_mode=GridUpdateMode.NO_UPDATE,
        fit_columns_on_grid_load=False, columns_auto_size_mode=2,
        allow_unsafe_jscode=True, enable_enterprise_modules=True,
        custom_css={
            ".ag-root-wrapper":       {"background-color": "#FFFFFF !important"},
            ".ag-header":             {"background-color": "#4A4A4A !important"},
            ".ag-header-row":         {"background-color": "#4A4A4A !important"},
            ".ag-header-cell":        {"background-color": "#4A4A4A !important"},
            ".ag-header-cell-label":  {"color": "#FFFFFF !important", "font-weight": "700"},
            ".ag-header-cell-text":   {"color": "#FFFFFF !important", "font-size": "13px !important", "font-weight": "700 !important"},
            ".ag-icon":               {"color": "#FFFFFF !important", "opacity": "1 !important"},
            ".ag-header-icon":        {"color": "#FFFFFF !important", "opacity": "1 !important"},
            # Headers coloridos por classe (só afetam a tabela que tem essas colunas)
            ".ag-header-cell.hdr-alta":  {"background-color": "#90EE90 !important"},
            ".ag-header-cell.hdr-alta .ag-header-cell-text": {"color": "#1A1A1A !important"},
            ".ag-header-cell.hdr-alta .ag-icon": {"color": "#1A1A1A !important"},
            ".ag-header-cell.hdr-sup":   {"background-color": "#87CEFF !important"},
            ".ag-header-cell.hdr-sup .ag-header-cell-text": {"color": "#1A1A1A !important"},
            ".ag-header-cell.hdr-sup .ag-icon": {"color": "#1A1A1A !important"},
            ".ag-header-cell.hdr-comp":  {"background-color": "#FFFF00 !important"},
            ".ag-header-cell.hdr-comp .ag-header-cell-text": {"color": "#1A1A1A !important"},
            ".ag-header-cell.hdr-comp .ag-icon": {"color": "#1A1A1A !important"},
            ".ag-header-cell.hdr-restr": {"background-color": "#FF0000 !important"},
            ".ag-header-cell.hdr-restr .ag-header-cell-text": {"color": "#FFFFFF !important"},
            ".ag-header-cell.hdr-restr .ag-icon": {"color": "#FFFFFF !important"},
            ".ag-header-cell[col-id='Alta Performance']": {"background-color": "#90EE90 !important"},
            ".ag-header-cell[col-id='Alta Performance'] .ag-header-cell-text": {"color": "#1A1A1A !important"},
            ".ag-header-cell[col-id='Superior']": {"background-color": "#87CEFF !important"},
            ".ag-header-cell[col-id='Superior'] .ag-header-cell-text": {"color": "#1A1A1A !important"},
            ".ag-header-cell[col-id='Competitivo']": {"background-color": "#FFFF00 !important"},
            ".ag-header-cell[col-id='Competitivo'] .ag-header-cell-text": {"color": "#1A1A1A !important"},
            ".ag-header-cell[col-id='Restrito']": {"background-color": "#FF0000 !important"},
            ".ag-header-cell[col-id='Restrito'] .ag-header-cell-text": {"color": "#FFFFFF !important"},
            ".ag-row":                {"background-color": "#FFFFFF !important", "color": "#1A1A1A !important", "font-size": "13px !important"},
            ".ag-row-odd":            {"background-color": "#F8FAF9 !important"},
            ".ag-cell":               {"color": "#1A1A1A !important", "font-size": "13px !important"},
        },
        theme="alpine", use_container_width=True, key=key,
    )


# ── Tabela HTML com header colorido por classe (100% confiável) ───────────────
def tabela_resumo_html(df: pd.DataFrame):
    """Tabela estática em HTML. Colunas cujo nome é uma classe ganham o header
    na cor da classe; as demais ficam no cinza padrão do painel."""
    cores_hdr = {
        "Alta Performance": ("#90EE90", "#1A1A1A"),
        "Superior":         ("#87CEFF", "#1A1A1A"),
        "Competitivo":      ("#FFFF00", "#1A1A1A"),
        "Restrito":         ("#FF0000", "#FFFFFF"),
    }
    ths = []
    for c in df.columns:
        bg, fg = cores_hdr.get(c, ("#4A4A4A", "#FFFFFF"))
        align = "left" if c == df.columns[0] else "center"
        ths.append(
            f'<th style="background:{bg};color:{fg};padding:9px 12px;'
            f'text-align:{align};font-size:13px;font-weight:700;'
            f'border-right:1px solid rgba(255,255,255,0.15);white-space:nowrap;">{c}</th>')

    trs = []
    for i, (_, row) in enumerate(df.iterrows()):
        bg_row = "#FFFFFF" if i % 2 == 0 else "#F8FAF9"
        tds = []
        for c in df.columns:
            v = row[c]
            txt = "" if pd.isna(v) else str(v)
            align = "left" if c == df.columns[0] else "center"
            peso = "700" if c == df.columns[0] else "400"
            tds.append(
                f'<td style="background:{bg_row};color:#1A1A1A;padding:8px 12px;'
                f'text-align:{align};font-size:13px;font-weight:{peso};'
                f'border-top:1px solid #EEEEEE;">{txt}</td>')
        trs.append("<tr>" + "".join(tds) + "</tr>")

    html = (
        '<table style="border-collapse:collapse;width:100%;border:1px solid #E5E7EB;'
        'border-radius:8px;overflow:hidden;font-family:Helvetica Neue,sans-serif;'
        'box-shadow:0 1px 4px rgba(0,0,0,0.06);">'
        '<thead><tr>' + "".join(ths) + '</tr></thead>'
        '<tbody>' + "".join(trs) + '</tbody></table>')
    st.markdown(html, unsafe_allow_html=True)


# ── Export Excel colorido pela classe ─────────────────────────────────────────
def to_excel(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    df = df.reset_index(drop=True)
    cols_drop = [c for c in df.columns
                 if str(c).startswith(("_", ":")) or "auto_unique_id" in str(c)]
    df = df.drop(columns=cols_drop).copy()

    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    fundo_xl = {k: v.lstrip("#") for k, v in COR_FUNDO.items()}
    fonte_xl = {k: v.lstrip("#") for k, v in COR_TEXTO.items()}

    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=str(col))
        if col in fundo_xl:                       # header de coluna de classe → cor da classe
            c.fill = PatternFill("solid", start_color=fundo_xl[col])
            c.font = Font(bold=True, name="Arial", size=10, color=fonte_xl.get(col, "1A1A1A"))
        else:
            c.fill = PatternFill("solid", start_color="4A4A4A")
            c.font = Font(bold=True, name="Arial", size=10, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = brd
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = max(14, len(str(col)) + 4)
    ws.row_dimensions[1].height = 28

    col_classe = df.columns.tolist().index("Classe") if "Classe" in df.columns else None

    for ri, row in enumerate(df.itertuples(index=False), start=2):
        classe_val = str(row[col_classe]) if col_classe is not None else "—"
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = brd
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="Arial", size=10)
            if col_classe is not None and (ci - 1) == col_classe:
                cell.fill = PatternFill("solid", start_color=fundo_xl.get(classe_val, "FFFFFF"))
                cell.font = Font(name="Arial", size=10, bold=True,
                                 color=fonte_xl.get(classe_val, "1A1A1A"))
    wb.save(buf)
    return buf.getvalue()


# ── Donut de classes ──────────────────────────────────────────────────────────
def donut_classes(contagem: dict) -> go.Figure:
    labels = ["Alta Performance", "Superior", "Competitivo", "Restrito"]
    valores = [contagem.get(l, 0) for l in labels]
    cores = [COR_FUNDO[l] for l in labels]
    fig = go.Figure(go.Pie(
        labels=labels, values=valores, hole=0.62,
        marker=dict(colors=cores, line=dict(color="#FFFFFF", width=2)),
        textinfo="value", textfont=dict(size=17, color="#1A1A1A"),
        sort=False,
    ))
    total = sum(valores)
    fig.update_layout(
        template="plotly_white",
        paper_bgcolor="#FFFFFF", plot_bgcolor="#FFFFFF",
        font=dict(family="Helvetica Neue, sans-serif", color="#000000", size=15),
        showlegend=True,
        legend=dict(orientation="h", yanchor="bottom", y=-0.18, x=0.5, xanchor="center",
                    font=dict(color="#1A1A1A", size=14)),
        margin=dict(t=10, b=10, l=10, r=10), height=320,
        annotations=[dict(text=f"<b>{total}</b><br>checks", x=0.5, y=0.5,
                          font=dict(size=22, color="#1A1A1A"), showarrow=False)],
    )
    return fig
